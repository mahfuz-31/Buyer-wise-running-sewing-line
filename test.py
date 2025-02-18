import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

file_path = "Hourly Report.html"
with open(file_path, "r", encoding="utf-8") as file:
    html_file = file.read()
html_content = BeautifulSoup(html_file, "html.parser")
rows = []
for row in html_content.find_all('tr'):
    row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
    rows.append(row_data)
floor_names = ['JAL-1', 'JAL-2', 'JAL-3', 'FFL-1', 'FFL-2', 'FFL-3', 'JFL-1', 'JFL-2', 'JKL-1', 'JKL-2', 'JKL-3', 'JKL-4', 'JKL-5', 'DBL', 'MFL', 'MFL-1', 'MFL-2',
'MFL-3', 'MFL-4', 'FFL2-1', 'FFL2-2', 'FFL2-3', 'FFL2-4', 'FFL2-5', 'JKL-U2-1', 'JKL-U2-2', 'JKL-U2-3', 'JKL-U2-4', 'JKL-U2-5']
result = pd.DataFrame()
result['Unit'] = None
result['Line'] = None
result['Buyer'] = None
lines = []
buyers = []
from typing import Dict
production: Dict[str, int] = {}
total_production = 0
index = 0
for row in rows:
    if len(row) > 8 and row[1] in floor_names:
        buyer = row[3].split('::')[0].split()[1]
        if buyer not in buyers:
            buyers.append(buyer)
        if buyer in production.keys():
            production[buyer] += int(row[len(row) - 2])
        else:
            production[buyer] = int(row[len(row) - 2])
        total_production += int(row[len(row) - 2])
        line = str(row[1]) + str(row[2]) + buyer
        if line not in lines:
            lines.append(line)
            result.loc[index, 'Unit'] = row[1]
            result.loc[index, 'Line'] = row[2]
            result.loc[index, 'Buyer'] = buyer
            index += 1
print('Total Line Running:', len(lines))

wb = load_workbook('Template.xlsx')
ws = wb['unit line buyer']
ws_prod = wb['Production_2']
for row in ws.iter_rows(min_row=2, max_col=3):
    for cell in row:
        cell.value = ''

for index, row in result.iterrows():
    u_idx = 'A' + str(index + 2)
    l_idx = 'B' + str(index + 2)
    b_idx = 'C' + str(index + 2)
    ws[u_idx] = row['Unit']
    ws[l_idx] = row['Line']
    ws[b_idx] = row['Buyer']
for i in range(100):
    b_idx = 'A' + str(i + 2)
    q_idx = 'C' + str(i + 2)
    ws_prod[b_idx] = ''
    ws_prod[q_idx] = ''
i = 2
for key, value in production.items():
    b_idx = 'A' + str(i)
    q_idx = 'C' + str(i)
    ws_prod[b_idx] = key
    ws_prod[q_idx] = int(value)
    i += 1
print("Total Production: ", total_production)
wb.save('Template.xlsx')